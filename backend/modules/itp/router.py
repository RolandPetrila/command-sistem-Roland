"""
API endpoints for ITP (Inspectie Tehnica Periodica) module.

Endpoints:
  GET    /api/itp/inspections                          — list all (paginated, searchable)
  POST   /api/itp/inspections                          — create new inspection
  GET    /api/itp/inspections/{id}                     — get single inspection
  PUT    /api/itp/inspections/{id}                     — update inspection
  DELETE /api/itp/inspections/{id}                     — delete inspection
  GET    /api/itp/vehicle/{plate}/history              — all inspections for a plate
  GET    /api/itp/rejection-reasons                    — standard ITP rejection reasons list
  POST   /api/itp/inspections/{id}/create-invoice      — pre-filled invoice data from inspection
  POST   /api/itp/import                               — import CSV/Excel file (with duplicate detection)
  GET    /api/itp/stats/overview                        — total, admis/respins, avg price, this month
  GET    /api/itp/stats/monthly                         — inspections per month (bar chart)
  GET    /api/itp/stats/brands                          — top brands (pie chart)
  GET    /api/itp/stats/revenue                         — monthly revenue (line chart)
  GET    /api/itp/stats/fuel-types                      — distribution by fuel type
  GET    /api/itp/stats/inspectors                      — stats per inspector
  GET    /api/itp/expiring                              — vehicles with ITP expiring soon
  GET    /api/itp/export/csv                            — export all as CSV
  GET    /api/itp/export/excel                          — export all as Excel
  GET    /api/itp/followup/due-soon                     — R4-24: vehicles due for re-inspection soon
  PUT    /api/itp/appointments/{id}/mark-showup         — R4-26: mark no-show / showed up
  GET    /api/itp/stats/noshow-rate                     — R4-26: no-show rate statistics
"""

from __future__ import annotations

import csv
import io
import logging
import os
import uuid
from datetime import date, timedelta
from pathlib import Path
from typing import Optional

import httpx
from fastapi import APIRouter, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse

from app.core.activity_log import log_activity
from app.db.database import get_db

from .imports import parse_csv, parse_excel
from .models import (
    ALLOWED_IMAGE_TYPES,
    ITP_PHOTOS_DIR,
    MAX_PHOTO_SIZE,
    MAX_PHOTOS_PER_INSPECTION,
    STANDARD_REJECTION_REASONS,
    AppointmentCreate,
    AppointmentUpdate,
    InspectionCreate,
    InspectionUpdate,
    MarkShowupRequest,
    row_dict,
    validate_appointment_transition,
)
from .statistics import (
    compute_followup_due_soon,
    compute_noshow_rate,
    compute_statistics_combined,
    compute_stats_brands,
    compute_stats_fuel_types,
    compute_stats_inspectors,
    compute_stats_monthly,
    compute_stats_overview,
    compute_stats_revenue,
)

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/itp", tags=["itp"])


# ────────── CRUD ──────────

@router.get("/inspections")
async def list_inspections(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: str = Query("", description="Search by plate number or owner name"),
    page: Optional[int] = Query(None, ge=1, description="Page number (alternative to skip)"),
    per_page: Optional[int] = Query(None, ge=1, le=500, description="Items per page (alternative to limit)"),
):
    """List all inspections with pagination and optional search.

    Accepts either skip/limit or page/per_page. If page is provided,
    it takes precedence and calculates skip automatically.
    """
    # If page/per_page provided, convert to skip/limit
    if page is not None:
        per_page = per_page or limit
        skip = (page - 1) * per_page
        limit = per_page

    async with get_db() as db:
        if search:
            pattern = f"%{search}%"
            cursor = await db.execute(
                """SELECT COUNT(*) FROM itp_inspections
                   WHERE plate_number LIKE ? OR owner_name LIKE ?""",
                (pattern, pattern),
            )
            total = (await cursor.fetchone())[0]

            cursor = await db.execute(
                """SELECT * FROM itp_inspections
                   WHERE plate_number LIKE ? OR owner_name LIKE ?
                   ORDER BY inspection_date DESC
                   LIMIT ? OFFSET ?""",
                (pattern, pattern, limit, skip),
            )
        else:
            cursor = await db.execute("SELECT COUNT(*) FROM itp_inspections")
            total = (await cursor.fetchone())[0]

            cursor = await db.execute(
                """SELECT * FROM itp_inspections
                   ORDER BY inspection_date DESC
                   LIMIT ? OFFSET ?""",
                (limit, skip),
            )

        rows = await cursor.fetchall()
        items = [row_dict(row) for row in rows]

    total_pages = (total + limit - 1) // limit if limit > 0 else 1
    current_page = (skip // limit) + 1 if limit > 0 else 1

    return {
        "items": items,
        "total": total,
        "skip": skip,
        "limit": limit,
        "total_pages": total_pages,
        "page": current_page,
    }


@router.post("/inspections")
async def create_inspection(data: InspectionCreate):
    """Create a new ITP inspection record."""
    async with get_db() as db:
        cursor = await db.execute(
            """INSERT INTO itp_inspections
               (plate_number, vin, brand, model, year, fuel_type,
                owner_name, owner_phone, inspection_date, expiry_date,
                result, rejection_reasons, price, inspector_name, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                data.plate_number.upper().strip(),
                data.vin,
                data.brand,
                data.model,
                data.year,
                data.fuel_type,
                data.owner_name,
                data.owner_phone,
                data.inspection_date,
                data.expiry_date,
                data.result,
                data.rejection_reasons,
                data.price,
                data.inspector_name,
                data.notes,
            ),
        )
        await db.commit()
        new_id = cursor.lastrowid

        # 3-strike rule: count previous rejections for this plate
        blocked = False
        if data.result == "Respins":
            plate_upper = data.plate_number.upper().strip()
            cnt_cursor = await db.execute(
                "SELECT COUNT(*) FROM itp_inspections WHERE UPPER(plate_number) = ? AND result = 'Respins'",
                (plate_upper,),
            )
            cnt_row = await cnt_cursor.fetchone()
            rejection_count = cnt_row[0] if cnt_row else 0
            if rejection_count >= 3:
                blocked = True
                logger.warning(
                    "3-STRIKE BLOCK: plate %s has %d rejections (including current)",
                    plate_upper, rejection_count,
                )

    await log_activity(
        action="itp.create",
        summary=f"ITP creat: {data.plate_number} — {data.result}",
        details={"id": new_id, "plate": data.plate_number, "result": data.result},
    )
    response = {"id": new_id, "message": "Inspectie creata cu succes"}
    if blocked:
        response["blocked"] = True
        response["blocked_message"] = (
            f"ATENTIE: {data.plate_number.upper().strip()} a acumulat 3 sau mai multe respingeri. "
            "Verificati vehiculul cu atentie inainte de urmatoarea inspectie."
        )
    return response


@router.get("/inspections/{inspection_id}")
async def get_inspection(inspection_id: int):
    """Get a single inspection by ID."""
    async with get_db() as db:
        cursor = await db.execute(
            "SELECT * FROM itp_inspections WHERE id = ?", (inspection_id,)
        )
        row = await cursor.fetchone()

    if not row:
        raise HTTPException(404, "Inspectia nu a fost gasita")
    return row_dict(row)


@router.put("/inspections/{inspection_id}")
async def update_inspection(inspection_id: int, data: InspectionUpdate):
    """Update an existing inspection."""
    # Build dynamic update
    fields = []
    values = []
    update_data = data.model_dump(exclude_unset=True)

    if not update_data:
        raise HTTPException(400, "Niciun camp de actualizat")

    for key, value in update_data.items():
        if key == "plate_number" and value:
            value = value.upper().strip()
        fields.append(f"{key} = ?")
        values.append(value)

    values.append(inspection_id)

    async with get_db() as db:
        cursor = await db.execute(
            f"UPDATE itp_inspections SET {', '.join(fields)} WHERE id = ?",
            values,
        )
        await db.commit()

        if cursor.rowcount == 0:
            raise HTTPException(404, "Inspectia nu a fost gasita")

    await log_activity(
        action="itp.update",
        summary=f"ITP actualizat: ID {inspection_id}",
        details={"id": inspection_id, "fields": list(update_data.keys())},
    )
    return {"message": "Inspectie actualizata cu succes"}


@router.delete("/inspections/{inspection_id}")
async def delete_inspection(inspection_id: int):
    """Delete an inspection."""
    async with get_db() as db:
        cursor = await db.execute(
            "DELETE FROM itp_inspections WHERE id = ?", (inspection_id,)
        )
        await db.commit()

        if cursor.rowcount == 0:
            raise HTTPException(404, "Inspectia nu a fost gasita")

    await log_activity(
        action="itp.delete",
        summary=f"ITP sters: ID {inspection_id}",
        details={"id": inspection_id},
    )
    return {"message": "Inspectie stearsa cu succes"}


# ────────── Vehicle History ──────────

@router.get("/vehicle/{plate}/history")
async def vehicle_history(plate: str):
    """All inspections for a specific plate number, ordered by date DESC."""
    plate_upper = plate.strip().upper()
    if not plate_upper:
        raise HTTPException(400, "Numarul de inmatriculare este obligatoriu")

    async with get_db() as db:
        cursor = await db.execute(
            """SELECT * FROM itp_inspections
               WHERE UPPER(plate_number) = ?
               ORDER BY inspection_date DESC""",
            (plate_upper,),
        )
        rows = await cursor.fetchall()

    items = [row_dict(row) for row in rows]
    total = len(items)
    admis = sum(1 for item in items if item.get("result") == "admis")
    respins = sum(1 for item in items if item.get("result") == "respins")
    pass_rate = round((admis / total * 100), 1) if total > 0 else 0

    return {
        "plate_number": plate_upper,
        "total": total,
        "admis": admis,
        "respins": respins,
        "pass_rate": pass_rate,
        "inspections": items,
    }


# ────────── Rejection Reasons ──────────

@router.get("/rejection-reasons")
async def get_rejection_reasons():
    """Standard ITP rejection reasons list."""
    return STANDARD_REJECTION_REASONS


# ────────── Generate Invoice from Inspection ──────────

@router.post("/inspections/{inspection_id}/create-invoice")
async def create_invoice_from_inspection(inspection_id: int):
    """Return pre-filled invoice data from an inspection record.

    Also stores the link back: updates the inspection with linked_invoice_id
    so the relationship is bidirectional (ITP -> Invoice).
    """
    async with get_db() as db:
        # Ensure linked_invoice_id column exists (ALTER TABLE IF NOT EXISTS pattern)
        try:
            await db.execute(
                "ALTER TABLE itp_inspections ADD COLUMN linked_invoice_id INTEGER"
            )
            await db.commit()
        except Exception:
            pass  # Column already exists

        cursor = await db.execute(
            "SELECT * FROM itp_inspections WHERE id = ?", (inspection_id,)
        )
        row = await cursor.fetchone()

    if not row:
        raise HTTPException(404, "Inspectia nu a fost gasita")

    inspection = row_dict(row)
    plate = inspection.get("plate_number", "")
    owner = inspection.get("owner_name", "")
    price = inspection.get("price", 0) or 0
    result = inspection.get("result", "")
    insp_date = inspection.get("inspection_date", "")

    invoice_data = {
        "client_name": owner or f"Proprietar {plate}",
        "client_phone": inspection.get("owner_phone", ""),
        "date": insp_date,
        "items": [
            {
                "description": f"Inspectie ITP - {plate} ({result})",
                "quantity": 1,
                "unit": "buc",
                "price": price,
                "total": price,
            }
        ],
        "total": price,
        "notes": f"Vehicul: {inspection.get('brand', '')} {inspection.get('model', '')} ({inspection.get('year', '')}) - {plate}",
        "source": "itp",
        "source_id": inspection_id,
        "linked_inspection_id": inspection_id,
    }

    await log_activity(
        action="itp.create_invoice",
        summary=f"Generare factura din ITP: {plate} — {price} RON",
        details={"inspection_id": inspection_id, "plate": plate, "price": price},
    )

    return invoice_data


@router.put("/inspections/{inspection_id}/link-invoice")
async def link_invoice_to_inspection(inspection_id: int, invoice_id: int):
    """Store the bidirectional link: set linked_invoice_id on the ITP inspection.

    Called after the invoice is actually created from the pre-filled data,
    so the inspection record knows which invoice was generated from it.
    """
    async with get_db() as db:
        # Ensure column exists
        try:
            await db.execute(
                "ALTER TABLE itp_inspections ADD COLUMN linked_invoice_id INTEGER"
            )
            await db.commit()
        except Exception:
            pass  # Column already exists

        cursor = await db.execute(
            "SELECT id FROM itp_inspections WHERE id = ?", (inspection_id,)
        )
        if not await cursor.fetchone():
            raise HTTPException(404, "Inspectia nu a fost gasita")

        await db.execute(
            "UPDATE itp_inspections SET linked_invoice_id = ? WHERE id = ?",
            (invoice_id, inspection_id),
        )
        await db.commit()

    await log_activity(
        action="itp.link_invoice",
        summary=f"ITP #{inspection_id} linkuit la factura #{invoice_id}",
        details={"inspection_id": inspection_id, "invoice_id": invoice_id},
    )
    return {
        "message": f"Inspectia #{inspection_id} linkuita la factura #{invoice_id}.",
        "inspection_id": inspection_id,
        "invoice_id": invoice_id,
    }


# ────────── Import CSV/Excel ──────────

@router.post("/import")
async def import_inspections(file: UploadFile = File(...)):
    """Import inspections from CSV or Excel file."""
    filename = (file.filename or "").lower()
    content = await file.read()

    try:
        if filename.endswith((".xlsx", ".xls")):
            rows = parse_excel(content)
        elif filename.endswith(".csv") or "csv" in (file.content_type or ""):
            rows = parse_csv(content)
        else:
            raise HTTPException(
                400,
                "Format nesuportat. Acceptate: CSV, XLSX",
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Eroare la parsarea fisierului: {e}")

    imported = 0
    skipped = 0
    duplicates = []
    errors = []

    async with get_db() as db:
        for i, row in enumerate(rows, 1):
            try:
                plate = (row.get("plate_number") or "").strip().upper()
                if not plate:
                    skipped += 1
                    continue

                insp_date = row.get("inspection_date", "")
                exp_date = row.get("expiry_date", "")
                result = (row.get("result") or "admis").strip().lower()

                if result not in ("admis", "respins"):
                    result = "admis"

                price = 0
                try:
                    price = float(row.get("price", 0) or 0)
                except (ValueError, TypeError):
                    price = 0

                year = None
                try:
                    year = int(row.get("year", 0) or 0)
                    if year < 1900:
                        year = None
                except (ValueError, TypeError):
                    year = None

                # Duplicate detection: same plate + same date
                if insp_date:
                    dup_cursor = await db.execute(
                        """SELECT id FROM itp_inspections
                           WHERE UPPER(plate_number) = ? AND inspection_date = ?
                           LIMIT 1""",
                        (plate, insp_date),
                    )
                    existing = await dup_cursor.fetchone()
                    if existing:
                        duplicates.append({
                            "row": i,
                            "plate_number": plate,
                            "inspection_date": insp_date,
                            "existing_id": existing[0],
                        })
                        continue

                await db.execute(
                    """INSERT INTO itp_inspections
                       (plate_number, brand, model, year, fuel_type,
                        owner_name, inspection_date, expiry_date,
                        result, price)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (
                        plate,
                        row.get("brand", ""),
                        row.get("model", ""),
                        year,
                        row.get("fuel_type", ""),
                        row.get("owner_name", ""),
                        insp_date,
                        exp_date,
                        result,
                        price,
                    ),
                )
                imported += 1
            except Exception as e:
                errors.append(f"Rand {i}: {e}")

        await db.commit()

    await log_activity(
        action="itp.import",
        summary=f"Import ITP: {imported} importate, {len(duplicates)} duplicate, {skipped} sarite, {len(errors)} erori",
        details={
            "file": file.filename,
            "imported": imported,
            "duplicates": len(duplicates),
            "skipped": skipped,
        },
    )

    return {
        "imported": imported,
        "skipped": skipped,
        "duplicates": duplicates[:50],  # Limit duplicate list
        "errors": errors[:20],  # Limit error list
        "message": f"{imported} inspectii importate cu succes"
        + (f", {len(duplicates)} duplicate sarite" if duplicates else ""),
    }


# ────────── Statistics ──────────

@router.get("/statistics")
async def statistics_combined():
    """Combined statistics endpoint for frontend dashboard."""
    return await compute_statistics_combined()


@router.get("/stats/overview")
async def stats_overview():
    """Overview statistics: total, admis/respins ratio, avg price, this month."""
    return await compute_stats_overview()


@router.get("/stats/monthly")
async def stats_monthly(year: int = Query(default=None)):
    """Inspections per month for a given year (default: current year)."""
    return await compute_stats_monthly(year)


@router.get("/stats/brands")
async def stats_brands():
    """Top car brands by inspection count."""
    return await compute_stats_brands()


@router.get("/stats/revenue")
async def stats_revenue(year: int = Query(default=None)):
    """Monthly revenue for a given year."""
    return await compute_stats_revenue(year)


@router.get("/stats/fuel-types")
async def stats_fuel_types():
    """Distribution by fuel type."""
    return await compute_stats_fuel_types()


@router.get("/stats/inspectors")
async def stats_inspectors():
    """Statistics per inspector: total, admis, respins, rate, revenue."""
    return await compute_stats_inspectors()


# ────────── Expiring ──────────

@router.get("/expiring")
async def expiring_inspections(days: int = Query(30, ge=1, le=365)):
    """List vehicles with ITP expiring within N days."""
    today = date.today().isoformat()
    future = (date.today() + timedelta(days=days)).isoformat()

    async with get_db() as db:
        cursor = await db.execute(
            """SELECT id, plate_number, brand, model, owner_name, owner_phone,
                      expiry_date, inspection_date
               FROM itp_inspections
               WHERE expiry_date >= ? AND expiry_date <= ?
               ORDER BY expiry_date ASC""",
            (today, future),
        )
        rows = await cursor.fetchall()

    result = []
    today_date = date.today()
    for row in rows:
        item = row_dict(row)
        try:
            exp = date.fromisoformat(item["expiry_date"])
            item["days_remaining"] = (exp - today_date).days
        except (ValueError, TypeError):
            item["days_remaining"] = None
        result.append(item)

    return result


# ────────── Export ──────────

@router.get("/export/csv")
async def export_csv():
    """Export all inspections as CSV."""
    async with get_db() as db:
        cursor = await db.execute(
            "SELECT * FROM itp_inspections ORDER BY inspection_date DESC"
        )
        rows = await cursor.fetchall()
        items = [row_dict(row) for row in rows]

    if not items:
        raise HTTPException(404, "Nu exista inspectii de exportat")

    output = io.StringIO()
    headers = [
        "id", "plate_number", "vin", "brand", "model", "year", "fuel_type",
        "owner_name", "owner_phone", "inspection_date", "expiry_date",
        "result", "rejection_reasons", "price", "inspector_name", "notes",
        "created_at",
    ]
    writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for item in items:
        writer.writerow(item)

    csv_bytes = output.getvalue().encode("utf-8-sig")

    await log_activity(
        action="itp.export",
        summary=f"Export CSV: {len(items)} inspectii",
        details={"format": "csv", "count": len(items)},
    )

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="itp_inspectii.csv"'},
    )


@router.get("/export/excel")
async def export_excel():
    """Export all inspections as Excel (XLSX)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    async with get_db() as db:
        cursor = await db.execute(
            "SELECT * FROM itp_inspections ORDER BY inspection_date DESC"
        )
        rows = await cursor.fetchall()
        items = [row_dict(row) for row in rows]

    if not items:
        raise HTTPException(404, "Nu exista inspectii de exportat")

    wb = Workbook()
    ws = wb.active
    ws.title = "Inspectii ITP"

    headers = [
        "ID", "Nr. Inmatriculare", "VIN", "Marca", "Model", "An", "Combustibil",
        "Proprietar", "Telefon", "Data ITP", "Data Expirare",
        "Rezultat", "Motiv Respingere", "Pret (RON)", "Inspector", "Note",
        "Data Creare",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    field_keys = [
        "id", "plate_number", "vin", "brand", "model", "year", "fuel_type",
        "owner_name", "owner_phone", "inspection_date", "expiry_date",
        "result", "rejection_reasons", "price", "inspector_name", "notes",
        "created_at",
    ]
    admis_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    respins_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, item in enumerate(items, 2):
        for col_idx, key in enumerate(field_keys, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))
            # Color-code result column
            if key == "result":
                if item.get("result") == "admis":
                    cell.fill = admis_fill
                elif item.get("result") == "respins":
                    cell.fill = respins_fill

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    await log_activity(
        action="itp.export",
        summary=f"Export Excel: {len(items)} inspectii",
        details={"format": "xlsx", "count": len(items)},
    )

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="itp_inspectii.xlsx"'},
    )


# ────────── F5: Appointments / Calendar ──────────

@router.get("/appointments")
async def list_appointments(
    date_from: str = Query(None),
    date_to: str = Query(None),
    status: str = Query(None),
):
    """List appointments with optional date range and status filter."""
    conditions = []
    params = []
    if date_from:
        conditions.append("scheduled_date >= ?")
        params.append(date_from)
    if date_to:
        conditions.append("scheduled_date <= ?")
        params.append(date_to)
    if status:
        conditions.append("status = ?")
        params.append(status)
    where = ("WHERE " + " AND ".join(conditions)) if conditions else ""
    async with get_db() as db:
        cursor = await db.execute(
            f"SELECT * FROM itp_appointments {where} ORDER BY scheduled_date ASC, scheduled_time ASC",
            params,
        )
        rows = await cursor.fetchall()
    return [row_dict(r) for r in rows]


@router.post("/appointments", status_code=201)
async def create_appointment(data: AppointmentCreate, force: bool = Query(False)):
    """Create a new ITP appointment. Checks for time conflicts unless force=True."""
    plate_upper = data.plate_number.upper().strip()
    conflict = None

    async with get_db() as db:
        # Check for overlapping appointments on the same date
        cursor = await db.execute(
            """SELECT id, plate_number, scheduled_time, duration_min
               FROM itp_appointments
               WHERE scheduled_date = ? AND status != 'cancelled'""",
            (data.scheduled_date,),
        )
        existing_appts = await cursor.fetchall()

        # Parse new appointment time range
        try:
            new_start_parts = data.scheduled_time.split(":")
            new_start_min = int(new_start_parts[0]) * 60 + int(new_start_parts[1])
            new_end_min = new_start_min + data.duration_min
        except (ValueError, IndexError):
            new_start_min = 0
            new_end_min = data.duration_min

        for appt in existing_appts:
            appt_dict = row_dict(appt)
            try:
                ex_parts = appt_dict["scheduled_time"].split(":")
                ex_start = int(ex_parts[0]) * 60 + int(ex_parts[1])
                ex_end = ex_start + (appt_dict["duration_min"] or 30)
            except (ValueError, IndexError):
                continue

            # Overlap check: two intervals [a,b) and [c,d) overlap if a < d and c < b
            if new_start_min < ex_end and ex_start < new_end_min:
                conflict = {
                    "existing_id": appt_dict["id"],
                    "existing_plate": appt_dict["plate_number"],
                    "existing_time": appt_dict["scheduled_time"],
                    "existing_duration": appt_dict["duration_min"],
                }
                break

        if conflict and not force:
            return {
                "warning": "Conflict de programare detectat",
                "conflict": conflict,
                "message": f"Exista deja o programare la {conflict['existing_time']} "
                           f"({conflict['existing_plate']}) in aceeasi zi. "
                           f"Folositi force=true pentru a crea oricum.",
                "created": False,
            }

        cursor = await db.execute(
            """INSERT INTO itp_appointments
               (plate_number, owner_name, owner_phone, scheduled_date, scheduled_time, duration_min, notes)
               VALUES (?, ?, ?, ?, ?, ?, ?)""",
            (plate_upper, data.owner_name, data.owner_phone,
             data.scheduled_date, data.scheduled_time, data.duration_min, data.notes),
        )
        await db.commit()
        appt_id = cursor.lastrowid

    await log_activity(
        action="itp.appointment_create",
        summary=f"Programare ITP: {plate_upper} pe {data.scheduled_date} la {data.scheduled_time}"
                + (" (conflict fortat)" if conflict else ""),
        details={"id": appt_id, "plate": plate_upper, "date": data.scheduled_date,
                 "conflict": conflict},
    )

    result = {"id": appt_id, "message": "Programare creata cu succes.", "created": True}
    if conflict:
        result["warning"] = "Programare creata cu conflict de timp"
        result["conflict"] = conflict
    return result


@router.put("/appointments/{appt_id}")
async def update_appointment(appt_id: int, data: AppointmentUpdate):
    """Update an appointment."""
    updates = data.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(400, "Niciun camp de actualizat")
    if "plate_number" in updates and updates["plate_number"]:
        updates["plate_number"] = updates["plate_number"].upper().strip()
    async with get_db() as db:
        # Validate status transition if status is being changed
        if "status" in updates:
            row = await db.execute(
                "SELECT status FROM itp_appointments WHERE id = ?", (appt_id,)
            )
            existing = await row.fetchone()
            if existing is None:
                raise HTTPException(404, "Programarea nu a fost gasita")
            validate_appointment_transition(existing["status"], updates["status"])
        fields = ", ".join(f"{k} = ?" for k in updates)
        values = list(updates.values()) + [appt_id]
        cursor = await db.execute(
            f"UPDATE itp_appointments SET {fields} WHERE id = ?", values
        )
        await db.commit()
        if cursor.rowcount == 0:
            raise HTTPException(404, "Programarea nu a fost gasita")
    return {"message": "Programare actualizata cu succes."}


@router.delete("/appointments/{appt_id}")
async def delete_appointment(appt_id: int):
    """Delete an appointment."""
    async with get_db() as db:
        cursor = await db.execute(
            "DELETE FROM itp_appointments WHERE id = ?", (appt_id,)
        )
        await db.commit()
        if cursor.rowcount == 0:
            raise HTTPException(404, "Programarea nu a fost gasita")
    return {"message": "Programare stearsa cu succes."}


@router.put("/appointments/{appt_id}/complete")
async def complete_appointment(appt_id: int, inspection_id: int = None):
    """Mark appointment as completed, optionally linking to an inspection."""
    async with get_db() as db:
        row = await db.execute(
            "SELECT status FROM itp_appointments WHERE id = ?", (appt_id,)
        )
        existing = await row.fetchone()
        if existing is None:
            raise HTTPException(404, "Programarea nu a fost gasita")
        validate_appointment_transition(existing["status"], "completed")
        cursor = await db.execute(
            "UPDATE itp_appointments SET status = 'completed', inspection_id = ? WHERE id = ?",
            (inspection_id, appt_id),
        )
        await db.commit()
        if cursor.rowcount == 0:
            raise HTTPException(404, "Programarea nu a fost gasita")
    return {"message": "Programare finalizata cu succes."}


# ────────── R4-24: Follow-up Alerts ──────────

@router.get("/followup/due-soon")
async def followup_due_soon(days: int = Query(30, ge=1, le=365)):
    """Vehicles with next inspection due within N days based on last inspection + 12 months."""
    return await compute_followup_due_soon(days)


# ────────── R4-26: No-show Tracking ──────────

@router.put("/appointments/{appt_id}/mark-showup")
async def mark_appointment_showup(appt_id: int, data: MarkShowupRequest):
    """Mark whether a client showed up for their appointment."""
    async with get_db() as db:
        # Ensure showed_up column exists
        try:
            await db.execute(
                "ALTER TABLE itp_appointments ADD COLUMN showed_up INTEGER"
            )
            await db.commit()
        except Exception:
            pass  # Column already exists

        row = await db.execute(
            "SELECT status FROM itp_appointments WHERE id = ?", (appt_id,)
        )
        existing = await row.fetchone()
        if existing is None:
            raise HTTPException(404, "Programarea nu a fost gasita")
        new_status = "completed" if data.showed_up else "no_show"
        validate_appointment_transition(existing["status"], new_status)

        cursor = await db.execute(
            "UPDATE itp_appointments SET showed_up = ?, status = ? WHERE id = ?",
            (1 if data.showed_up else 0, new_status, appt_id),
        )
        await db.commit()
        if cursor.rowcount == 0:
            raise HTTPException(404, "Programarea nu a fost gasita")

    label = "prezent" if data.showed_up else "neprezentare"
    await log_activity(
        action="itp.appointment_showup",
        summary=f"Programare #{appt_id}: {label}",
        details={"id": appt_id, "showed_up": data.showed_up},
    )
    return {"message": f"Programare marcata ca {label}."}


@router.get("/stats/noshow-rate")
async def stats_noshow_rate():
    """No-show statistics for appointments."""
    return await compute_noshow_rate()


# ────────── Photos ──────────

@router.post("/inspections/{inspection_id}/photos")
async def upload_inspection_photo(inspection_id: int, file: UploadFile = File(...)):
    """Upload a photo for an inspection. Max 5MB, images only, max 5 per inspection."""
    # Validate inspection exists
    async with get_db() as db:
        cursor = await db.execute(
            "SELECT id FROM itp_inspections WHERE id = ?", (inspection_id,)
        )
        if not await cursor.fetchone():
            raise HTTPException(404, "Inspectia nu a fost gasita")

        # Check current photo count
        cursor = await db.execute(
            "SELECT COUNT(*) FROM itp_photos WHERE inspection_id = ?",
            (inspection_id,),
        )
        count = (await cursor.fetchone())[0]
        if count >= MAX_PHOTOS_PER_INSPECTION:
            raise HTTPException(
                400,
                f"Limita de {MAX_PHOTOS_PER_INSPECTION} fotografii per inspectie a fost atinsa.",
            )

    # Validate content type
    content_type = file.content_type or ""
    if content_type not in ALLOWED_IMAGE_TYPES:
        raise HTTPException(
            400,
            f"Tip fisier neacceptat: {content_type}. Acceptate: JPEG, PNG, WebP, GIF.",
        )

    # Read and validate size
    content = await file.read()
    if len(content) > MAX_PHOTO_SIZE:
        raise HTTPException(400, f"Fisierul depaseste limita de {MAX_PHOTO_SIZE // (1024*1024)} MB.")

    # Generate unique filename
    ext = Path(file.filename or "photo.jpg").suffix.lower() or ".jpg"
    unique_name = f"{uuid.uuid4().hex}{ext}"

    # Create directory for this inspection
    photo_dir = ITP_PHOTOS_DIR / str(inspection_id)
    photo_dir.mkdir(parents=True, exist_ok=True)
    filepath = photo_dir / unique_name

    # Resize if width > 1920px using PIL
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(content))
        if img.width > 1920:
            ratio = 1920 / img.width
            new_size = (1920, int(img.height * ratio))
            img = img.resize(new_size, Image.LANCZOS)
        img.save(str(filepath), quality=85)
    except ImportError:
        # PIL not available — save raw
        filepath.write_bytes(content)
    except Exception as exc:
        # Fallback: save raw on any PIL error
        logger.warning("PIL resize failed, saving raw: %s", exc)
        filepath.write_bytes(content)

    # Save to database
    async with get_db() as db:
        cursor = await db.execute(
            "INSERT INTO itp_photos (inspection_id, filename, filepath) VALUES (?, ?, ?)",
            (inspection_id, file.filename or unique_name, str(filepath)),
        )
        await db.commit()
        photo_id = cursor.lastrowid

    await log_activity(
        action="itp.photo_upload",
        summary=f"Foto ITP incarcata: inspectie #{inspection_id} — {file.filename}",
        details={"inspection_id": inspection_id, "photo_id": photo_id},
    )

    return {"id": photo_id, "filename": file.filename or unique_name, "path": str(filepath)}


@router.get("/inspections/{inspection_id}/photos")
async def list_inspection_photos(inspection_id: int):
    """List all photos for an inspection."""
    async with get_db() as db:
        cursor = await db.execute(
            "SELECT id FROM itp_inspections WHERE id = ?", (inspection_id,)
        )
        if not await cursor.fetchone():
            raise HTTPException(404, "Inspectia nu a fost gasita")

        cursor = await db.execute(
            "SELECT id, inspection_id, filename, filepath, created_at "
            "FROM itp_photos WHERE inspection_id = ? ORDER BY created_at DESC",
            (inspection_id,),
        )
        rows = await cursor.fetchall()

    return [row_dict(row) for row in rows]


@router.delete("/inspections/{inspection_id}/photos/{photo_id}")
async def delete_inspection_photo(inspection_id: int, photo_id: int):
    """Delete a photo from an inspection."""
    async with get_db() as db:
        cursor = await db.execute(
            "SELECT id, filepath, filename FROM itp_photos WHERE id = ? AND inspection_id = ?",
            (photo_id, inspection_id),
        )
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(404, "Fotografia nu a fost gasita")

        filepath = Path(row["filepath"])
        filename = row["filename"]

        # Delete from DB
        await db.execute("DELETE FROM itp_photos WHERE id = ?", (photo_id,))
        await db.commit()

    # Delete file from disk
    try:
        if filepath.exists():
            filepath.unlink()
    except OSError as exc:
        logger.warning("Nu s-a putut sterge fisierul foto: %s", exc)

    await log_activity(
        action="itp.photo_delete",
        summary=f"Foto ITP stearsa: inspectie #{inspection_id} — {filename}",
        details={"inspection_id": inspection_id, "photo_id": photo_id},
    )

    return {"message": "Fotografia a fost stearsa cu succes."}


@router.get("/photos/serve/{photo_id}")
async def serve_photo(photo_id: int):
    """Serve a photo file by its ID."""
    async with get_db() as db:
        cursor = await db.execute(
            "SELECT filepath, filename FROM itp_photos WHERE id = ?", (photo_id,)
        )
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(404, "Fotografia nu a fost gasita")

    filepath = Path(row["filepath"])
    if not filepath.exists():
        raise HTTPException(404, "Fisierul foto nu mai exista pe disk")

    return FileResponse(
        str(filepath),
        filename=row["filename"],
        media_type="image/jpeg",
    )


# ────────── ITP Expiry Notification Cron ──────────

async def check_itp_expiring_notifications():
    """Check for ITP inspections expiring within 14 days and send Telegram notifications.

    Queries inspections where expiry_date is within 14 days and notified_expiry = 0.
    Sends a Telegram message for each and marks notified_expiry = 1.
    """
    bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
    chat_id = os.getenv("TELEGRAM_CHAT_ID")

    if not bot_token or not chat_id:
        logger.debug("Telegram credentials not set, skipping ITP expiry notifications")
        return

    today = date.today()
    threshold = (today + timedelta(days=14)).isoformat()
    today_str = today.isoformat()

    try:
        async with get_db() as db:
            # Ensure notified_expiry column exists
            try:
                await db.execute(
                    "ALTER TABLE itp_inspections ADD COLUMN notified_expiry INTEGER DEFAULT 0"
                )
                await db.commit()
            except Exception:
                pass  # Column already exists

            cursor = await db.execute(
                """SELECT id, plate_number, owner_name, brand, model, expiry_date
                   FROM itp_inspections
                   WHERE expiry_date >= ? AND expiry_date <= ?
                     AND (notified_expiry IS NULL OR notified_expiry = 0)
                   ORDER BY expiry_date ASC""",
                (today_str, threshold),
            )
            rows = await cursor.fetchall()

        if not rows:
            return

        notified_ids = []
        url = f"https://api.telegram.org/bot{bot_token}/sendMessage"

        async with httpx.AsyncClient(timeout=15.0) as client:
            for row in rows:
                inspection = row_dict(row)
                days_left = (date.fromisoformat(inspection["expiry_date"]) - today).days
                plate = inspection["plate_number"]
                owner = inspection.get("owner_name") or "N/A"
                vehicle = f"{inspection.get('brand') or ''} {inspection.get('model') or ''}".strip() or "N/A"

                message = (
                    f"🚗 ITP Expira in {days_left} zile!\n"
                    f"Nr: {plate}\n"
                    f"Proprietar: {owner}\n"
                    f"Vehicul: {vehicle}\n"
                    f"Data expirare: {inspection['expiry_date']}"
                )

                try:
                    resp = await client.post(url, json={
                        "chat_id": chat_id,
                        "text": message,
                        "parse_mode": "HTML",
                    })
                    if resp.status_code == 200:
                        notified_ids.append(inspection["id"])
                    else:
                        logger.warning(
                            "Telegram send failed for ITP #%d: HTTP %d",
                            inspection["id"], resp.status_code,
                        )
                except Exception as exc:
                    logger.warning("Telegram send error for ITP #%d: %s", inspection["id"], exc)

        # Mark as notified
        if notified_ids:
            async with get_db() as db:
                placeholders = ",".join("?" * len(notified_ids))
                await db.execute(
                    f"UPDATE itp_inspections SET notified_expiry = 1 WHERE id IN ({placeholders})",
                    notified_ids,
                )
                await db.commit()

            await log_activity(
                action="itp.expiry_notifications",
                summary=f"Notificari ITP expirare: {len(notified_ids)} trimise via Telegram",
                details={"notified_ids": notified_ids},
            )
            logger.info("ITP expiry notifications sent: %d", len(notified_ids))

    except Exception as exc:
        logger.error("Error checking ITP expiry notifications: %s", exc)
