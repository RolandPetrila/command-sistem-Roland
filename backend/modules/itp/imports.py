"""
CSV/Excel import parsing logic for the ITP module.

Pure functions — no route decorators, no database access.
"""

from __future__ import annotations

import csv
import io


def parse_csv(content: bytes) -> list[dict]:
    """Parse CSV content into list of dicts."""
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    return list(reader)


def parse_excel(content: bytes) -> list[dict]:
    """Parse Excel content into list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    # First row = headers
    headers_raw = next(rows_iter, None)
    if not headers_raw:
        return []

    headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}"
               for i, h in enumerate(headers_raw)]

    result = []
    for row in rows_iter:
        obj = {}
        for i, val in enumerate(row):
            key = headers[i] if i < len(headers) else f"col_{i}"
            if val is not None:
                obj[key] = str(val) if hasattr(val, "isoformat") else val
            else:
                obj[key] = ""
        result.append(obj)

    wb.close()
    return result
